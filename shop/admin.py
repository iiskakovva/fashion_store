from django.contrib import admin
from django.contrib.auth.admin import UserAdmin
from django.shortcuts import render
from django.http import HttpResponse
from django.urls import path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from .models import User, Category, Product, Size, Cart, CartItem, Order, OrderItem, Review
from django.utils import timezone

class CustomUserAdmin(UserAdmin):
    list_display = ('username', 'email', 'first_name', 'last_name', 'phone', 'is_staff')
    list_filter = ('is_staff', 'is_superuser', 'is_active')
    fieldsets = UserAdmin.fieldsets + (
        ('Дополнительная информация', {'fields': ('phone', 'address', 'avatar')}),
    )

class CategoryAdmin(admin.ModelAdmin):
    prepopulated_fields = {'slug': ('name',)}
    list_display = ('name', 'slug', 'created_at')
    search_fields = ('name',)

class SizeInline(admin.TabularInline):
    model = Size
    extra = 1

class ProductAdmin(admin.ModelAdmin):
    list_display = ('name', 'category', 'price', 'style', 'stock', 'available', 'created_at')
    list_filter = ('category', 'style', 'available')
    search_fields = ('name', 'description')
    prepopulated_fields = {'slug': ('name',)}
    list_editable = ('price', 'stock', 'available')
    inlines = [SizeInline]
    
    actions = ['export_products_excel']
    
    def export_products_excel(self, request, queryset):
        """Экспорт выбранных товаров в Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products Export"
        
        # Заголовки
        headers = ['Название', 'Категория', 'Цена', 'Стиль', 'Количество', 'Доступен', 'Дата создания']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Данные
        for row, product in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=product.name)
            ws.cell(row=row, column=2, value=product.category.name)
            ws.cell(row=row, column=3, value=float(product.price))
            ws.cell(row=row, column=4, value=product.get_style_display())
            ws.cell(row=row, column=5, value=product.stock)
            ws.cell(row=row, column=6, value="Да" if product.available else "Нет")
            ws.cell(row=row, column=7, value=product.created_at.strftime('%Y-%m-%d %H:%M'))
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="products_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response
    export_products_excel.short_description = "Экспортировать выбранные товары в Excel"

class OrderAdmin(admin.ModelAdmin):
    list_display = ('id', 'user', 'first_name', 'last_name', 'total_price', 'status', 'created_at')
    list_filter = ('status', 'created_at', 'city')
    search_fields = ('first_name', 'last_name', 'email', 'phone')
    list_editable = ('status',)
    readonly_fields = ('total_price',)
    
    actions = ['export_orders_excel']
    
    def export_orders_excel(self, request, queryset):
        """Экспорт выбранных заказов в Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders Export"
        
        headers = ['№ заказа', 'Пользователь', 'Имя', 'Фамилия', 'Email', 'Сумма', 'Статус', 'Город', 'Дата']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        for row, order in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=order.id)
            ws.cell(row=row, column=2, value=order.user.username)
            ws.cell(row=row, column=3, value=order.first_name)
            ws.cell(row=row, column=4, value=order.last_name)
            ws.cell(row=row, column=5, value=order.email)
            ws.cell(row=row, column=6, value=float(order.total_price))
            ws.cell(row=row, column=7, value=order.get_status_display())
            ws.cell(row=row, column=8, value=order.city)
            ws.cell(row=row, column=9, value=order.created_at.strftime('%Y-%m-%d %H:%M'))
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="orders_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response
    export_orders_excel.short_description = "Экспортировать выбранные заказы в Excel"

class ReviewAdmin(admin.ModelAdmin):
    list_display = ('product', 'user', 'rating', 'created_at')
    list_filter = ('rating', 'created_at')
    search_fields = ('product__name', 'user__username', 'comment')

# Регистрация моделей в админке
admin.site.register(User, CustomUserAdmin)
admin.site.register(Category, CategoryAdmin)
admin.site.register(Product, ProductAdmin)  # Только одна регистрация Product
admin.site.register(Order, OrderAdmin)
admin.site.register(Review, ReviewAdmin)