import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from django.utils import timezone

def export_to_excel(queryset, fields, filename_prefix):
    """
    Утилита для экспорта queryset в Excel файл
    """
    # Создаем книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{filename_prefix}"
    
    # Получаем человекочитаемые названия полей
    model = queryset.model
    headers = []
    for field_name in fields:
        if field_name == 'category':
            headers.append('Категория')
        elif field_name == 'user':
            headers.append('Пользователь')
        else:
            # Пытаемся получить verbose_name поля
            try:
                field = model._meta.get_field(field_name)
                headers.append(field.verbose_name.title())
            except:
                headers.append(field_name.replace('_', ' ').title())
    
    # Стилизация заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Данные
    for row, obj in enumerate(queryset, 2):
        for col, field_name in enumerate(fields, 1):
            value = getattr(obj, field_name, '')
            
            # Обработка специальных случаев
            if field_name == 'category' and hasattr(obj, 'category'):
                value = obj.category.name if obj.category else ''
            elif field_name == 'user' and hasattr(obj, 'user'):
                value = str(obj.user)
            elif hasattr(value, 'strftime'):  # Для дат
                value = value.strftime('%Y-%m-%d %H:%M')
            
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="left")
    
    # Автоподбор ширины колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Создаем response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response